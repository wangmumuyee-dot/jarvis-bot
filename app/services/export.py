from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Literal

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from app.security.sensitive import redact_sensitive_text
from app.storage.db import Database
from app.services.ledger import LedgerService, parse_ledger_text, _type_label

ExportScope = Literal["all", "month", "reimbursable"]


@dataclass(frozen=True)
class ExportResult:
    path: Path
    row_count: int
    scope: ExportScope
    redacted: bool = False


@dataclass(frozen=True)
class ImportResult:
    path: Path
    row_count: int


class LedgerExportService:
    def __init__(self, db: Database, export_dir: Path) -> None:
        self.db = db
        self.export_dir = export_dir

    def export(self, scope: ExportScope, now: datetime | None = None, *, redact_sensitive: bool = False) -> ExportResult:
        now = now or datetime.now()
        rows = self._rows(scope, now)
        self.export_dir.mkdir(parents=True, exist_ok=True)
        suffix = "-redacted" if redact_sensitive else ""
        path = self.export_dir / f"ledger-{scope}{suffix}-{now.strftime('%Y%m%d-%H%M%S')}.xlsx"
        self._write_workbook(path, rows, redact_sensitive=redact_sensitive)
        return ExportResult(path=path, row_count=len(rows), scope=scope, redacted=redact_sensitive)

    def _rows(self, scope: ExportScope, now: datetime) -> list[dict[str, object]]:
        where = ""
        params: list[object] = []
        if scope == "month":
            start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            end = start.replace(year=start.year + 1, month=1) if start.month == 12 else start.replace(month=start.month + 1)
            where = "WHERE occurred_at >= ? AND occurred_at < ?"
            params = [start.isoformat(timespec="seconds"), end.isoformat(timespec="seconds")]
        elif scope == "reimbursable":
            where = "WHERE reimbursable = 1 AND reimbursement_status = 'pending'"

        with self.db.connect() as conn:
            records = conn.execute(
                f"""
                SELECT
                    ledger_entries.id,
                    ledger_entries.entry_type,
                    ledger_entries.amount,
                    ledger_entries.currency,
                    ledger_entries.category,
                    ledger_entries.note,
                    ledger_entries.occurred_at,
                    ledger_entries.reimbursable,
                    ledger_entries.reimbursement_status,
                    ledger_entries.source_text,
                    ledger_entries.created_at,
                    COALESCE(ledger_books.name, '日常账本') AS book,
                    COALESCE(ledger_accounts.name, '默认账户') AS account,
                    GROUP_CONCAT(ledger_tags.name, ',') AS tags
                FROM ledger_entries
                LEFT JOIN ledger_books ON ledger_books.id = ledger_entries.book_id
                LEFT JOIN ledger_accounts ON ledger_accounts.id = ledger_entries.account_id
                LEFT JOIN ledger_entry_tags ON ledger_entry_tags.entry_id = ledger_entries.id
                LEFT JOIN ledger_tags ON ledger_tags.id = ledger_entry_tags.tag_id
                {where}
                GROUP BY ledger_entries.id
                ORDER BY ledger_entries.occurred_at DESC, ledger_entries.id DESC
                """,
                params,
            ).fetchall()
        return [dict(record) for record in records]

    def _write_workbook(self, path: Path, rows: list[dict[str, object]], *, redact_sensitive: bool = False) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "账本流水"
        headers = [
            "ID",
            "账本",
            "账户",
            "类型",
            "金额",
            "币种",
            "分类",
            "备注",
            "标签",
            "发生时间",
            "是否待报销",
            "报销状态",
            "原始输入",
            "创建时间",
        ]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for row in rows:
            note = redact_sensitive_text(row["note"]) if redact_sensitive else row["note"]
            source_text = redact_sensitive_text(row["source_text"]) if redact_sensitive else row["source_text"]
            sheet.append(
                [
                    row["id"],
                    row["book"],
                    row["account"],
                    _type_label(row["entry_type"]),
                    row["amount"],
                    row["currency"],
                    row["category"],
                    note,
                    row["tags"] or "",
                    row["occurred_at"],
                    "是" if row["reimbursable"] else "否",
                    row["reimbursement_status"],
                    source_text,
                    row["created_at"],
                ]
            )

        for column in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 10), 36)
        workbook.save(path)

    def import_xlsx(self, path: Path, now: datetime | None = None) -> ImportResult:
        workbook = load_workbook(path)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return ImportResult(path=path, row_count=0)

        headers = [str(value or "").strip() for value in rows[0]]
        created = 0
        ledger = LedgerService(self.db)
        for raw in rows[1:]:
            row = {headers[index]: raw[index] for index in range(min(len(headers), len(raw)))}
            amount = row.get("金额")
            note = row.get("备注") or row.get("原始输入") or row.get("分类") or ""
            if amount in {None, ""}:
                continue
            text_parts = [
                str(note),
                str(amount),
                f"用{row.get('账户')}" if row.get("账户") else "",
                f"记到{row.get('账本')}" if row.get("账本") else "",
                " ".join(f"#{tag.strip()}" for tag in str(row.get("标签") or "").split(",") if tag.strip()),
            ]
            draft = parse_ledger_text(" ".join(part for part in text_parts if part), now=now)
            if not draft:
                continue
            ledger.create(draft)
            created += 1
        return ImportResult(path=path, row_count=created)


def parse_export_scope(text: str) -> ExportScope | None:
    if "导入" in text:
        return None
    if "导出" not in text and "Excel" not in text and "excel" not in text and "账单" not in text:
        return None
    if "待报销" in text:
        return "reimbursable"
    if "本月" in text or "这个月" in text:
        return "month"
    return "all"


def handle_export_text(text: str, service: LedgerExportService) -> str | None:
    scope = parse_export_scope(text)
    if not scope:
        return None
    redacted = any(token in text for token in ["脱敏", "隐私", "安全"])
    result = service.export(scope, redact_sensitive=redacted)
    privacy = "脱敏" if result.redacted else ""
    return f"已导出{privacy}账本 Excel：{result.path}，共 {result.row_count} 条记录。"


def verify_xlsx(path: Path) -> bool:
    workbook = load_workbook(path)
    return bool(workbook.sheetnames)
