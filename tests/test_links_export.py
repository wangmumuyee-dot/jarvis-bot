from __future__ import annotations

import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.ai.summarize import SummaryService, summarize_locally
import app.services.knowledge as knowledge_module
from app.services.export import LedgerExportService, handle_export_text, verify_xlsx
from app.services.knowledge import KnowledgeService, handle_knowledge_text
from app.services.ledger import LedgerService, parse_ledger_text
from app.storage.db import Database
from app.utils.links import FetchedPage, extract_page_text


HTML_PAGE = """
        <html>
          <head><title>AI Agent 测试文章</title></head>
          <body>
            <article>
              <h1>AI Agent 测试文章</h1>
              <p>人工智能正在改变个人知识管理。</p>
              <p>个人机器人可以帮助用户整理链接、生成摘要，并写入 Obsidian。</p>
              <p>下一步可以接入更稳定的网页正文提取。</p>
            </article>
          </body>
        </html>
        """


class LinkAndExportTest(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.db = Database(self.root / "jarvis.db")
        self.db.init()

    def tearDown(self) -> None:
        self.tmp.cleanup()

    def test_fetch_public_page_extracts_title_and_text(self) -> None:
        page = extract_page_text("https://example.com/article", HTML_PAGE)
        self.assertEqual(page.title, "AI Agent 测试文章")
        self.assertIn("人工智能正在改变个人知识管理", page.text)

    def test_capture_link_writes_markdown_with_source_url(self) -> None:
        summary_service = SummaryService(
            client=type(
                "LocalSummaryClient",
                (),
                {
                    "configured": lambda _self: False,
                    "summarize": lambda _self, **kwargs: summarize_locally(**kwargs),
                },
            )()
        )
        knowledge = KnowledgeService(self.db, self.root / "vault", summary_service)
        original_fetch = knowledge_module.fetch_public_page
        knowledge_module.fetch_public_page = lambda url: FetchedPage(
            url=url,
            title="AI Agent 测试文章",
            text=extract_page_text(url, HTML_PAGE).text,
        )
        try:
            reply = handle_knowledge_text("总结这个链接 https://example.com/article", knowledge)
        finally:
            knowledge_module.fetch_public_page = original_fetch
        assert reply is not None
        self.assertIn("已写入 Obsidian 笔记", reply)
        files = list((self.root / "vault").rglob("*.md"))
        self.assertEqual(len(files), 1)
        content = files[0].read_text(encoding="utf-8")
        self.assertIn("source_type: \"link\"", content)
        self.assertIn("source_url:", content)
        self.assertIn("人工智能", content)

    def test_export_month_ledger_to_valid_xlsx(self) -> None:
        ledger = LedgerService(self.db)
        for text in ["今天午饭 38", "打车 26", "打车 48，待报销"]:
            draft = parse_ledger_text(text, now=datetime.now())
            assert draft is not None
            ledger.create(draft)

        export_service = LedgerExportService(self.db, self.root / "exports")
        reply = handle_export_text("导出本月账单", export_service)
        assert reply is not None
        self.assertIn("已导出账本 Excel", reply)
        files = list((self.root / "exports").glob("*.xlsx"))
        self.assertEqual(len(files), 1)
        self.assertTrue(verify_xlsx(files[0]))
        workbook = load_workbook(files[0])
        headers = [cell.value for cell in workbook.active[1]]
        self.assertIn("账本", headers)
        self.assertIn("账户", headers)
        self.assertIn("标签", headers)

    def test_import_ledger_xlsx(self) -> None:
        path = self.root / "import.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["账本", "账户", "金额", "分类", "备注", "标签"])
        sheet.append(["旅行账本", "招行信用卡", 88, "餐饮", "午饭", "出差"])
        workbook.save(path)

        export_service = LedgerExportService(self.db, self.root / "exports")
        result = export_service.import_xlsx(path, now=datetime(2026, 6, 8, 12, 0, 0))
        self.assertEqual(result.row_count, 1)

        ledger = LedgerService(self.db)
        query = ledger.query("这个月餐饮花了多少", now=datetime(2026, 6, 8, 12, 0, 0))
        assert query is not None
        self.assertEqual(query.total, 88)


if __name__ == "__main__":
    unittest.main()
